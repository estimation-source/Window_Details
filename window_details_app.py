from __future__ import annotations

import io
import math
import os
import re
import sys
from dataclasses import asdict, dataclass, field
from typing import Any, Dict, List, Optional, Tuple

import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from PIL import Image
import streamlit as st

# ============================================================
# 1. Streamlit Page Config
# ============================================================
st.set_page_config(
    page_title="WIN-SQUARE | Universal Window Details Engine",
    layout="wide",
    page_icon="🪟",
    initial_sidebar_state="expanded"
)

# ============================================================
# 2. FIX CSS: Header चालू ठेवून Sidebar Toggle Button Visible ठेवणे
# ============================================================
st.markdown("""
    <style>
    header[data-testid="stHeader"] {
        z-index: 99999 !important;
        background: transparent !important;
    }

    button[data-testid="stSidebarCollapsedControl"],
    button[data-testid="stSidebarNavCollapseButton"] {
        display: flex !important;
        visibility: visible !important;
        opacity: 1 !important;
        background-color: #FF4B4B !important;
        color: white !important;
        border-radius: 8px !important;
        position: fixed !important;
        top: 12px !important;
        left: 12px !important;
        z-index: 999999 !important;
        box-shadow: 0px 3px 8px rgba(0,0,0,0.3) !important;
    }

    button[data-testid="stSidebarCollapsedControl"] svg,
    button[data-testid="stSidebarNavCollapseButton"] svg {
        fill: white !important;
        color: white !important;
        width: 22px !important;
        height: 22px !important;
    }

    [data-testid="stStatusWidget"],
    #MainMenu, 
    footer {
        display: none !important;
        visibility: hidden !important;
    }
    </style>
""", unsafe_allow_html=True)

# State Management
if "uploader_key" not in st.session_state:
    st.session_state["uploader_key"] = 0

# ============================================================
# 3. UI Layout & Fonts CSS
# ============================================================
st.markdown(
    """
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&display=swap');
    
    html, body, [class*="css"] {
        font-family: 'Inter', sans-serif;
        background-color: #f4f6f9;
        color: #334155;
    }

    .main .block-container {
        padding-top: 1.5rem;
        padding-bottom: 2rem;
        max-width: 98%;
    }

    [data-testid="stSidebar"] {
        background-color: #f1f5f9;
        border-right: 1px solid #e2e8f0;
    }
    
    .quick-guide-title {
        font-size: 15px;
        font-weight: 700;
        color: #0f172a;
        margin-top: 15px;
        margin-bottom: 12px;
    }
    
    .quick-guide-step {
        font-size: 13px;
        color: #475569;
        margin-bottom: 10px;
        line-height: 1.4;
    }

    .hero-container {
        background: #ffffff;
        border-radius: 16px;
        padding: 24px 30px;
        border: 1px solid #e2e8f0;
        box-shadow: 0 2px 4px rgba(0,0,0,0.02);
        margin-bottom: 24px;
        display: flex;
        justify-content: space-between;
        align-items: center;
    }

    .hero-title-text {
        font-size: 22px;
        font-weight: 800;
        color: #0f172a;
        margin: 0;
    }

    .hero-sub-text {
        font-size: 13px;
        color: #64748b;
        margin-top: 4px;
    }

    .step-title {
        font-size: 16px;
        font-weight: 700;
        color: #1e293b;
        margin-bottom: 12px;
        display: flex;
        align-items: center;
        gap: 8px;
    }

    .kpi-card-box {
        background: #ffffff;
        border: 1px solid #e2e8f0;
        border-radius: 10px;
        padding: 16px 20px;
        box-shadow: 0 1px 2px rgba(0,0,0,0.03);
    }

    .kpi-title-lbl {
        font-size: 11px;
        font-weight: 700;
        color: #64748b;
        text-transform: uppercase;
        letter-spacing: 0.5px;
    }

    .kpi-val-lbl {
        font-size: 24px;
        font-weight: 800;
        color: #0f172a;
        margin-top: 6px;
    }

    /* CUSTOM FLEX CONTAINER FOR CLOSE BUTTONS LIKE IMAGE 2 */
    .button-group-container {
        display: flex;
        align-items: center;
        gap: 10px !important;
        margin-top: 10px;
        margin-bottom: 15px;
    }

    /* PRIMARY BLUE BUTTON - COMPACT & NORMAL FONT WEIGHT */
    div.stButton > button[kind="primary"] {
        background-color: #2563eb !important;
        background: #2563eb !important;
        border: 1px solid #2563eb !important;
        color: #ffffff !important;
        border-radius: 6px !important;
        height: 38px !important;
        padding: 0 16px !important;
        box-shadow: 0 1px 2px rgba(37, 99, 235, 0.2) !important;
    }
    div.stButton > button[kind="primary"]:hover {
        background-color: #1d4ed8 !important;
        background: #1d4ed8 !important;
    }

    /* SECONDARY RED BUTTON - COMPACT & NORMAL FONT WEIGHT */
    div.stButton > button[kind="secondary"] {
        background-color: #dc2626 !important;
        background: #dc2626 !important;
        border: 1px solid #dc2626 !important;
        color: #ffffff !important;
        border-radius: 6px !important;
        height: 38px !important;
        padding: 0 16px !important;
        box-shadow: 0 1px 2px rgba(220, 38, 38, 0.2) !important;
    }
    div.stButton > button[kind="secondary"]:hover {
        background-color: #b91c1c !important;
        background: #b91c1c !important;
    }

    /* DOWNLOAD GREEN BUTTON - COMPACT & NORMAL FONT WEIGHT */
    div.stDownloadButton > button {
        background-color: #059669 !important;
        background: #059669 !important;
        border: 1px solid #059669 !important;
        color: #ffffff !important;
        border-radius: 6px !important;
        height: 38px !important;
        padding: 0 16px !important;
        box-shadow: 0 1px 2px rgba(5, 150, 105, 0.2) !important;
    }
    div.stDownloadButton > button:hover {
        background-color: #047857 !important;
        background: #047857 !important;
    }

    /* FORCE NORMAL WEIGHT (NOT BOLD) & WHITE TEXT */
    div.stButton > button p, div.stButton > button span,
    div.stDownloadButton > button p, div.stDownloadButton > button span {
        color: #ffffff !important;
        font-weight: 500 !important;
        font-size: 13px !important;
    }

    .stTabs [data-baseweb="tab-list"] {
        gap: 16px;
        border-bottom: 1px solid #e2e8f0;
    }

    .stTabs [data-baseweb="tab"] {
        height: 40px;
        white-space: pre;
        font-size: 13px;
        font-weight: 600;
        color: #64748b;
    }

    .stTabs [aria-selected="true"] {
        color: #2563eb !important;
        border-bottom: 2px solid #2563eb !important;
    }
    </style>
    """,
    unsafe_allow_html=True,
)


def get_image_path(filename):
    if hasattr(sys, "_MEIPASS"):
        return os.path.join(sys._MEIPASS, filename)
    return os.path.join(os.path.abspath("."), filename)


# =========================================================
# SIDEBAR
# =========================================================
with st.sidebar:
    logo_file = get_image_path("logo.png")
    if os.path.exists(logo_file):
        col_s1, col_s2, col_s3 = st.columns([1, 2, 1])
        with col_s2:
            st.image(Image.open(logo_file), width=110)
    else:
        st.markdown("<h2 style='text-align: center; color:#1e293b;'><b>win square</b></h2>", unsafe_allow_html=True)
    
    st.markdown("---")
    st.markdown("<div class='quick-guide-title'>💡 Quick Guide</div>", unsafe_allow_html=True)
    st.markdown(
        """
        <div class='quick-guide-step'><b>1.</b> Upload multi-sheet Excel BOQ files.</div>
        <div class='quick-guide-step'><b>2.</b> Click on <b>Process Sheet</b>.</div>
        <div class='quick-guide-step'><b>3.</b> Review window records in table or dashboard tabs.</div>
        <div class='quick-guide-step'><b>4.</b> Click <b>Generate Windows Details Sheet</b>.</div>
        <div class='quick-guide-step'><b>5.</b> Download formatted Excel sheet with SQFT breakdown.</div>
        <div class='quick-guide-step'><b>6.</b> Use <b>Reset Data</b> to clear workspace.</div>
        """,
        unsafe_allow_html=True
    )


# =========================================================
# HEADER HERO BANNER
# =========================================================
st.markdown(
    """
    <div class="hero-container">
        <div>
            <div class="hero-title-text">Universal Window Details & Glass SQFT Engine</div>
            <div class="hero-sub-text">Supports Measurement Sheets, Quotation Sheets & Block Layouts</div>
        </div>
    </div>
    """,
    unsafe_allow_html=True,
)

# ============================================================
# Global Engine Constants & Parsing Logic (WINDOWS DETAILS)
# ============================================================

HEADER_SCAN_LIMIT = 200
HEADER_REMOVE_PATTERN = r"[^A-Z0-9]"

KEYWORDS = {
    "CODE": ["CODE", "WINDOW CODE", "LOCATION", "REF"],
    "WIDTH": [
        ["FWIDTH"], ["F", "WIDTH"], ["FRAME", "W"], ["WIDTH"], ["W"]
    ],
    "HEIGHT": [
        ["FHEIGHT"], ["F", "HEIGHT"], ["FRAME", "H"], ["HEIGHT"], ["H"]
    ],
    "QTY": ["QTY", "QUANTITY", "NO", "NOS"],
    "GLASS": [["GLASS"], ["GLASS", "SPEC"], ["REMARKS"], ["DESCRIPTION"]],
}


def standardize_glass_spec(val: str) -> str:
    if pd.isna(val) or not str(val).strip():
        return "NOT SPECIFIED"
    text = str(val).strip()
    if text.lower() == "nan" or not text:
        return "NOT SPECIFIED"
    return re.sub(r"\s+", " ", text).strip()


@dataclass(slots=True)
class HeaderInfo:
    row_index: int
    code_col: Optional[int] = None
    width_col: Optional[int] = None
    height_col: Optional[int] = None
    qty_col: Optional[int] = None
    glass_col: Optional[int] = None
    columns: Dict[str, Optional[int]] = field(default_factory=dict)


@dataclass(slots=True)
class HeaderBlock:
    header: HeaderInfo
    start_row: int
    end_row: int


@dataclass(slots=True)
class WindowRecord:
    WindowCode: str
    Width: int
    Height: int
    Qty: int
    GlassType: str
    SourceFile: str
    SheetName: str


def normalize_header(text: Any) -> str:
    if pd.isna(text):
        return ""
    text = str(text).upper().strip()
    return re.sub(HEADER_REMOVE_PATTERN, "", text)


def normalize_header_row(row: pd.Series) -> List[str]:
    return [normalize_header(val) for val in row.tolist()]


def contains_keywords(text: str, keyword_groups: List[Any]) -> bool:
    if not text:
        return False
    text = normalize_header(text)
    text = re.sub(r"[^A-Z0-9]", "", text)

    for group in keyword_groups:
        if isinstance(group, str):
            group = [group]
        matched = True
        for keyword in group:
            key = re.sub(r"[^A-Z0-9]", "", normalize_header(keyword))
            if key not in text:
                matched = False
                break
        if matched:
            return True
    return False


def detect_column(header_row: List[str], keyword_groups: List[Any]) -> Optional[int]:
    for index, value in enumerate(header_row):
        text = normalize_header(value)
        text = re.sub(r"[^A-Z0-9]", "", text)
        for group in keyword_groups:
            if isinstance(group, str):
                group = [group]
            matched = True
            for keyword in group:
                key = re.sub(r"[^A-Z0-9]", "", normalize_header(keyword))
                if key not in text:
                    matched = False
                    break
            if matched:
                return index
    return None


def detect_header_columns(header_row: pd.Series) -> Dict[str, Optional[int]]:
    normalized = normalize_header_row(header_row)
    columns = {
        "code": detect_column(normalized, KEYWORDS["CODE"]),
        "width": detect_column(normalized, KEYWORDS["WIDTH"]),
        "height": detect_column(normalized, KEYWORDS["HEIGHT"]),
        "qty": detect_column(normalized, KEYWORDS["QTY"]),
        "glass": detect_column(normalized, KEYWORDS["GLASS"]),
    }
    return columns


def is_business_header(row: pd.Series) -> bool:
    normalized = normalize_header_row(row)
    has_code = False
    has_qty = False
    has_dim = False

    for value in normalized:
        if contains_keywords(value, KEYWORDS["CODE"]):
            has_code = True
        if contains_keywords(value, KEYWORDS["QTY"]):
            has_qty = True
        if contains_keywords(value, KEYWORDS["WIDTH"]) or contains_keywords(value, KEYWORDS["HEIGHT"]):
            has_dim = True

    return has_code and (has_qty or has_dim)


def find_header_blocks(dataframe: pd.DataFrame) -> List[HeaderInfo]:
    headers: List[HeaderInfo] = []
    rows = min(len(dataframe), HEADER_SCAN_LIMIT)

    for row_number in range(rows):
        row = dataframe.iloc[row_number]
        if not is_business_header(row):
            continue

        columns = detect_header_columns(row)
        header = HeaderInfo(
            row_index=row_number,
            code_col=columns["code"],
            width_col=columns["width"],
            height_col=columns["height"],
            qty_col=columns["qty"],
            glass_col=columns["glass"],
            columns=columns,
        )
        headers.append(header)

    return headers


def build_header_blocks(dataframe: pd.DataFrame, headers: List[HeaderInfo]) -> List[HeaderBlock]:
    blocks: List[HeaderBlock] = []
    if not headers:
        return blocks

    headers = sorted(headers, key=lambda h: h.row_index)
    for i, header in enumerate(headers):
        start = header.row_index + 1
        end = (
            len(dataframe) - 1
            if i == len(headers) - 1
            else headers[i + 1].row_index - 1
        )
        blocks.append(HeaderBlock(header=header, start_row=start, end_row=end))

    return blocks


def safe_numeric(value: Any) -> Optional[int]:
    if pd.isna(value):
        return None
    try:
        val = float(value)
        if val <= 0:
            return None
        return int(math.floor(val + 0.5))
    except Exception:
        return None


def parse_header_block(dataframe: pd.DataFrame, block: HeaderBlock, source_file: str, sheet_name: str) -> List[WindowRecord]:
    records: List[WindowRecord] = []
    
    for row_no in range(block.start_row, block.end_row + 1):
        row = dataframe.iloc[row_no]
        
        # Code extraction
        code_val = row.iloc[block.header.code_col] if block.header.code_col is not None and block.header.code_col < len(row) else None
        if pd.isna(code_val) or not str(code_val).strip():
            continue
        
        code_str = str(code_val).strip().replace(".0", "")
        if code_str.upper() in ["CODE", "TOTAL", "SUBTOTAL"]:
            continue

        width_val = safe_numeric(row.iloc[block.header.width_col]) if block.header.width_col is not None and block.header.width_col < len(row) else None
        height_val = safe_numeric(row.iloc[block.header.height_col]) if block.header.height_col is not None and block.header.height_col < len(row) else None
        qty_val = safe_numeric(row.iloc[block.header.qty_col]) if block.header.qty_col is not None and block.header.qty_col < len(row) else 1
        
        glass_val = row.iloc[block.header.glass_col] if block.header.glass_col is not None and block.header.glass_col < len(row) else "NOT SPECIFIED"
        glass_str = standardize_glass_spec(glass_val)

        if width_val and height_val:
            records.append(
                WindowRecord(
                    WindowCode=code_str,
                    Width=width_val,
                    Height=height_val,
                    Qty=qty_val if qty_val else 1,
                    GlassType=glass_str,
                    SourceFile=source_file,
                    SheetName=sheet_name,
                )
            )

    return records


def parse_business_sheet(dataframe: pd.DataFrame, source_file: str, sheet_name: str) -> List[WindowRecord]:
    headers = find_header_blocks(dataframe)
    blocks = build_header_blocks(dataframe, headers)
    all_records: List[WindowRecord] = []

    for block in blocks:
        all_records.extend(parse_header_block(dataframe, block, source_file, sheet_name))

    return all_records


def load_excel_with_calculated_values(file) -> Dict[str, pd.DataFrame]:
    file_bytes = io.BytesIO(file.read())
    file.seek(0)
    wb = openpyxl.load_workbook(file_bytes, data_only=True)
    workbook_dict = {}

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        data = sheet.values
        cols = next(data, None)
        if cols is None:
            continue

        data_rows = list(data)
        if cols:
            data_rows.insert(0, cols)

        df = pd.DataFrame(data_rows)
        workbook_dict[sheet_name] = df

    return workbook_dict


def process_uploaded_files(uploaded_files) -> pd.DataFrame:
    all_records = []

    for file in uploaded_files:
        try:
            workbook_dict = load_excel_with_calculated_values(file)
            for sheet_name, df in workbook_dict.items():
                records = parse_business_sheet(df, file.name, sheet_name)
                all_records.extend(records)
        except Exception as e:
            st.error(f"Error processing file {file.name}: {e}")

    return pd.DataFrame([asdict(r) for r in all_records]).reset_index(drop=True) if all_records else pd.DataFrame()


# ============================================================
# STEP 1: FILE UPLOAD SECTION
# ============================================================
st.markdown("<div class='step-title'>📁 Step 1: Upload BOQ Excel Files</div>", unsafe_allow_html=True)

uploaded_files = st.file_uploader(
    "Upload BOQ Excel Files",
    type=["xlsx", "xlsm", "xls"],
    accept_multiple_files=True,
    label_visibility="collapsed",
    key=f"boq_uploader_{st.session_state['uploader_key']}"
)

# CLOSE BUTTON LAYOUT LIKE IMAGE 2
st.markdown("<div class='button-group-container'>", unsafe_allow_html=True)
btn_col1, btn_col2, _ = st.columns([0.20, 0.18, 0.62])

with btn_col1:
    btn_merge = st.button("🔗 Process Sheet", type="primary", use_container_width=False)

with btn_col2:
    btn_reset = st.button("🗑️ Reset Data", type="secondary", use_container_width=False)

st.markdown("</div>", unsafe_allow_html=True)

if btn_merge:
    if uploaded_files:
        with st.spinner("Processing & Extracting Window Records..."):
            df_merged = process_uploaded_files(uploaded_files)
            if not df_merged.empty:
                st.session_state["merged_df"] = df_merged
                st.toast(f"Successfully Extracted {len(df_merged)} Window Records!", icon="✅")
            else:
                st.error("⚠️ No valid window records found in uploaded file(s).")
    else:
        st.warning("Please upload Excel file(s) first!")

if btn_reset:
    for key in ["merged_df", "win_df_preview", "win_bytes", "win_generated"]:
        if key in st.session_state:
            del st.session_state[key]
    st.session_state["uploader_key"] += 1
    st.rerun()


# ============================================================
# EXTRACTED MASTER WINDOW RECORDS & DASHBOARD
# ============================================================
if "merged_df" in st.session_state:
    df_merged = st.session_state["merged_df"]

    st.markdown("<br>", unsafe_allow_html=True)
    st.markdown("<div class='step-title'>📋 Extracted Window Details & Master Records</div>", unsafe_allow_html=True)

    f_col1, f_col2 = st.columns([2, 2])
    with f_col1:
        search_query = st.text_input("🔍 Quick Search (Window Code / Glass Spec)", placeholder="Type to filter...")
    with f_col2:
        glass_types = ["ALL"] + sorted(list(df_merged["GlassType"].unique()))
        selected_glass = st.selectbox("Filter by Glass Spec", glass_types)

    filtered_df = df_merged.copy()
    if search_query:
        filtered_df = filtered_df[
            filtered_df["WindowCode"].str.contains(search_query, case=False, na=False) |
            filtered_df["GlassType"].str.contains(search_query, case=False, na=False)
        ]
    if selected_glass != "ALL":
        filtered_df = filtered_df[filtered_df["GlassType"] == selected_glass]

    filtered_display_df = filtered_df.copy()
    if "Sr. No." not in filtered_display_df.columns:
        filtered_display_df.insert(0, "Sr. No.", range(1, len(filtered_display_df) + 1))

    st.dataframe(filtered_display_df, use_container_width=True, height=260, hide_index=True)
    st.caption(f"Showing {len(filtered_df)} of {len(df_merged)} total extracted records")

    # ============================================================
    # STEP 2: DASHBOARD KPI CARDS & GENERATE BUTTON
    # ============================================================
    st.markdown("<br>", unsafe_allow_html=True)
    st.markdown("<div class='step-title'>⚡ Step 2: Dashboard Analytics & Processing</div>", unsafe_allow_html=True)

    if st.button("⚡ GENERATE WINDOWS DETAILS SHEET (EXCEL)", type="primary", use_container_width=False):
        with st.spinner("Calculating SQFT and preparing Dashboard Breakdown..."):
            df_win_preview = df_merged.copy()
            df_win_preview["SQFT"] = ((df_win_preview["Width"] * df_win_preview["Height"]) / 92903.04).round(6)
            df_win_preview["TTL SQFT"] = (df_win_preview["SQFT"] * df_win_preview["Qty"]).round(6)

            df_win_preview.insert(0, "Sr.No", range(1, len(df_win_preview) + 1))
            df_win_preview = df_win_preview.rename(
                columns={
                    "WindowCode": "WINDOW CODE",
                    "Width": "WIDTH",
                    "Height": "HEIGHT",
                    "Qty": "QTY",
                    "GlassType": "REMARKS",
                }
            )

            preview_cols = ["Sr.No", "WINDOW CODE", "WIDTH", "HEIGHT", "SQFT", "QTY", "TTL SQFT", "REMARKS"]
            df_win_preview = df_win_preview[preview_cols]
            st.session_state["win_df_preview"] = df_win_preview

            # OpenPyXL Sheet Processing
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "WINDOW DETAILS"
            ws.views.sheetView[0].showGridLines = True

            title_font = Font(name="Calibri", size=12, bold=True)
            header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            data_font = Font(name="Calibri", size=10)
            total_font = Font(name="Calibri", size=11, bold=True)

            header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
            total_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

            thin_border = Border(
                left=Side(style="thin", color="D9D9D9"),
                right=Side(style="thin", color="D9D9D9"),
                top=Side(style="thin", color="D9D9D9"),
                bottom=Side(style="thin", color="D9D9D9"),
            )
            thick_top_double_bottom = Border(
                top=Side(style="thin", color="000000"),
                bottom=Side(style="double", color="000000"),
            )

            ws.cell(row=1, column=1, value="WIN-SQUARE WINDOW DETAILS").font = title_font

            headers = ["Sr.No", "WINDOW CODE", "WIDTH", "HEIGHT", "SQFT", "QTY", "TTL SQFT", "REMARKS"]
            ws.append(headers)

            for c in range(1, len(headers) + 1):
                cell = ws.cell(row=2, column=c)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center" if c in [1, 6] else ("right" if c in [3, 4, 5, 7] else "left"))

            for idx, row in df_merged.iterrows():
                r_idx = idx + 3
                sqft_formula = f"=ROUND((C{r_idx}*D{r_idx})/92903.04, 6)"
                ttl_sqft_formula = f"=E{r_idx}*F{r_idx}"

                ws.append([
                    idx + 1, row["WindowCode"], row["Width"], row["Height"],
                    sqft_formula, row["Qty"], ttl_sqft_formula, row["GlassType"],
                ])

                for c in range(1, len(headers) + 1):
                    cell = ws.cell(row=r_idx, column=c)
                    cell.font = data_font
                    cell.border = thin_border
                    if c in [3, 4]:
                        cell.number_format = "0"
                        cell.alignment = Alignment(horizontal="right")
                    elif c == 5:
                        cell.number_format = "0.000000"
                        cell.alignment = Alignment(horizontal="right")
                    elif c == 6:
                        cell.number_format = "0"
                        cell.alignment = Alignment(horizontal="center")
                    elif c == 7:
                        cell.number_format = "0.000000"
                        cell.alignment = Alignment(horizontal="right")

            tot_row = len(df_merged) + 3
            ws.cell(row=tot_row, column=5, value="TOTAL").font = total_font
            ws.cell(row=tot_row, column=5).alignment = Alignment(horizontal="right")

            qty_sum = ws.cell(row=tot_row, column=6, value=f"=SUM(F3:F{tot_row-1})")
            qty_sum.font = total_font
            qty_sum.number_format = "0"
            qty_sum.alignment = Alignment(horizontal="center")

            ttl_sqft_sum = ws.cell(row=tot_row, column=7, value=f"=SUM(G3:G{tot_row-1})")
            ttl_sqft_sum.font = total_font
            ttl_sqft_sum.number_format = "0.000000"
            ttl_sqft_sum.alignment = Alignment(horizontal="right")

            for c in range(1, len(headers) + 1):
                cell = ws.cell(row=tot_row, column=c)
                cell.fill = total_fill
                cell.border = thick_top_double_bottom

            for col in ws.columns:
                max_len = max(len(str(cell.value or "")) for cell in col)
                col_letter = get_column_letter(col[0].column)
                ws.column_dimensions[col_letter].width = max(max_len + 3, 14)

            output = io.BytesIO()
            wb.save(output)
            st.session_state["win_bytes"] = output.getvalue()
            st.session_state["win_generated"] = True

    # Render KPI Cards & Live Dashboard Preview
    if st.session_state.get("win_generated"):
        st.markdown("<br>", unsafe_allow_html=True)
        
        win_df = st.session_state["win_df_preview"]
        tot_items = len(win_df)
        tot_qty = win_df["QTY"].sum()
        tot_area = win_df["TTL SQFT"].sum().round(2)

        k1, k2, k3 = st.columns(3)
        with k1:
            st.markdown(f"<div class='kpi-card-box'><div class='kpi-title-lbl'>TOTAL WINDOW TYPES</div><div class='kpi-val-lbl'>{tot_items}</div></div>", unsafe_allow_html=True)
        with k2:
            st.markdown(f"<div class='kpi-card-box'><div class='kpi-title-lbl'>TOTAL WINDOW QUANTITY</div><div class='kpi-val-lbl'>{tot_qty} Pcs</div></div>", unsafe_allow_html=True)
        with k3:
            st.markdown(f"<div class='kpi-card-box'><div class='kpi-title-lbl'>TOTAL GLASS AREA (SQFT)</div><div class='kpi-val-lbl'>{tot_area:,.2f} Sq.Ft</div></div>", unsafe_allow_html=True)
    
        st.markdown("<br>", unsafe_allow_html=True)

        tab1, tab2, tab3 = st.tabs([
            "📄 WINDOW DETAILS Live Preview", 
            "📊 OC Wise Summary (Windows & SQFT)", 
            "🧩 Glass Type Breakdown"
        ])

        with tab1:
            st.dataframe(win_df, use_container_width=True, height=350, hide_index=True)

        with tab2:
            # OC WISE SUMMARY
            df_merged_copy = df_merged.copy()
            df_merged_copy["Total_SQFT"] = ((df_merged_copy["Width"] * df_merged_copy["Height"]) / 92903.04) * df_merged_copy["Qty"]

            oc_summary = (
                df_merged_copy.groupby("SourceFile", as_index=False)
                .agg(
                    Qty=("Qty", "sum"),
                    Total_SQFT=("Total_SQFT", "sum")
                )
            )

            oc_summary["Total_SQFT"] = oc_summary["Total_SQFT"].round(2)
            oc_summary.columns = ["SourceFile (OC Name)", "Qty (Pcs)", "Total Glass SQFT"]
            
            if "Sr. No." not in oc_summary.columns:
                oc_summary.insert(0, "Sr. No.", range(1, len(oc_summary) + 1))

            st.dataframe(oc_summary, use_container_width=True, hide_index=True)

        with tab3:
            # GLASS TYPE BREAKDOWN
            df_glass_copy = df_merged.copy()
            df_glass_copy["Total_SQFT"] = ((df_glass_copy["Width"] * df_glass_copy["Height"]) / 92903.04) * df_glass_copy["Qty"]

            glass_breakdown = (
                df_glass_copy.groupby("GlassType", as_index=False)
                .agg(
                    Qty=("Qty", "sum"),
                    Total_SQFT=("Total_SQFT", "sum")
                )
                .sort_values(by="Qty", ascending=False)
            )

            glass_breakdown["Total_SQFT"] = glass_breakdown["Total_SQFT"].round(2)
            glass_breakdown.columns = ["Glass Type Specification", "Total Quantity (Pcs)", "Total Glass SQFT"]
            glass_breakdown.insert(0, "Sr. No.", range(1, len(glass_breakdown) + 1))

            st.dataframe(glass_breakdown, use_container_width=True, hide_index=True)

        st.markdown("<br>", unsafe_allow_html=True)
        st.success("✅ Windows Details Excel Sheet Ready! Complete with KPI Dashboard, OC summary, and Calibri styling.")
        
        st.download_button(
            label="📥 DOWNLOAD WINDOW DETAILS SHEET (.XLSX)",
            data=st.session_state["win_bytes"],
            file_name="WINDOWS_DETAILS_SHEET.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=False
        )
